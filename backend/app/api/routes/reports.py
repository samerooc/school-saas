"""
Reports & Analytics API - Module 5
Excel export for attendance, fees, results + dashboard stats
"""
import io
from datetime import date, datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func

from app.db.session import get_db
from app.core.security import require_roles
from app.models.models import Student, User, Class, Attendance, FeePayment, FeeStructure, Exam, Mark, Subject

router = APIRouter(prefix="/reports", tags=["reports"])

def _make_excel(sheets: dict) -> io.BytesIO:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    HFILL = PatternFill("solid", fgColor="4F46E5")
    HFONT = Font(color="FFFFFF", bold=True, size=11)
    AFILL = PatternFill("solid", fgColor="F0F0FF")
    BS = Side(style="thin", color="D1D5DB")
    CB = Border(left=BS, right=BS, top=BS, bottom=BS)

    for sname, data in sheets.items():
        ws = wb.create_sheet(title=sname[:31])
        headers, rows = data["headers"], data["rows"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill, c.font, c.border = HFILL, HFONT, CB
            c.alignment = Alignment(horizontal="center", vertical="center")
        for ri, row in enumerate(rows, 2):
            fill = AFILL if ri % 2 == 0 else None
            for col, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=col, value=val)
                c.border = CB
                c.alignment = Alignment(vertical="center")
                if fill: c.fill = fill
        for col in range(1, len(headers) + 1):
            ml = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, len(rows) + 2))
            ws.column_dimensions[get_column_letter(col)].width = min(ml + 4, 40)
        ws.row_dimensions[1].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/attendance/excel")
async def attendance_excel(
    from_date: date = Query(...), to_date: date = Query(...),
    class_id: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(require_roles("admin", "principal", "teacher")),
):
    school_id = payload["school_id"]
    if (to_date - from_date).days > 90:
        raise HTTPException(400, "Max 90 days range")

    import sqlalchemy
    query = (
        select(User.full_name, Student.roll_number, Class.name, Class.section,
               func.count(Attendance.id).label("total"),
               func.sum((Attendance.status == "present").cast(sqlalchemy.Integer)).label("present_count"))
        .join(Student, Student.id == Attendance.student_id)
        .join(User, User.id == Student.user_id, isouter=True)
        .join(Class, Class.id == Student.class_id, isouter=True)
        .where(Attendance.school_id == school_id, Attendance.date >= from_date, Attendance.date <= to_date)
        .group_by(User.full_name, Student.roll_number, Class.name, Class.section)
        .order_by(Class.name, Student.roll_number)
    )
    if class_id: query = query.where(Attendance.class_id == class_id)
    rows_data = (await db.execute(query)).all()

    rows = []
    for name, roll, cn, sec, total, present in rows_data:
        p, t = int(present or 0), int(total or 0)
        pct = round(p / t * 100, 1) if t > 0 else 0
        rows.append([name or "?", roll or "—", f"{cn or ''}{' '+sec if sec else ''}", t, p, t-p, f"{pct}%", "Low ⚠" if pct < 75 else "OK ✓"])

    buf = _make_excel({"Attendance": {"headers": ["Student","Roll No","Class","Total Days","Present","Absent","Percentage","Status"], "rows": rows}})
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Attendance_{from_date}_{to_date}.xlsx"})


@router.get("/fees/excel")
async def fees_excel(
    from_date: Optional[date] = Query(None), to_date: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(require_roles("admin", "principal")),
):
    school_id = payload["school_id"]
    query = (
        select(User.full_name, Student.admission_number, Class.name, Class.section,
               FeeStructure.fee_type, FeePayment.amount_paid, FeePayment.payment_method,
               FeePayment.receipt_number, FeePayment.payment_date)
        .join(Student, Student.id == FeePayment.student_id)
        .join(User, User.id == Student.user_id, isouter=True)
        .join(Class, Class.id == Student.class_id, isouter=True)
        .join(FeeStructure, FeeStructure.id == FeePayment.fee_structure_id, isouter=True)
        .where(FeePayment.school_id == school_id, FeePayment.status == "paid")
        .order_by(FeePayment.payment_date.desc())
    )
    if from_date: query = query.where(FeePayment.payment_date >= from_date)
    if to_date: query = query.where(FeePayment.payment_date <= to_date)
    rows_data = (await db.execute(query)).all()
    total = sum(float(r[5] or 0) for r in rows_data)
    rows = [[r[0] or "?", r[1] or "?", f"{r[2] or ''}{' '+r[3] if r[3] else ''}", r[4] or "?",
             f"₹{float(r[5] or 0):,.2f}", (r[6] or "").title(), r[7] or "?",
             r[8].strftime("%d %b %Y") if r[8] else "?"] for r in rows_data]
    rows.append(["","","","TOTAL", f"₹{total:,.2f}","","",""])
    buf = _make_excel({"Fee Collection": {"headers": ["Student","Adm. No","Class","Fee Type","Amount","Method","Receipt No","Date"], "rows": rows}})
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=FeeReport_{date.today()}.xlsx"})


@router.get("/dashboard/stats")
async def dashboard_stats(
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(require_roles("admin", "principal")),
):
    school_id = payload["school_id"]
    today = date.today()
    total_students = (await db.execute(select(func.count(Student.id)).where(Student.school_id == school_id, Student.is_active == True))).scalar() or 0
    total_staff = (await db.execute(select(func.count(User.id)).where(User.school_id == school_id, User.is_active == True, User.role.in_(["teacher", "principal", "staff"])))).scalar() or 0
    today_present = (await db.execute(select(func.count(Attendance.id)).where(Attendance.school_id == school_id, Attendance.date == today, Attendance.status == "present"))).scalar() or 0
    today_total = (await db.execute(select(func.count(Attendance.id)).where(Attendance.school_id == school_id, Attendance.date == today))).scalar() or 0
    month_start = today.replace(day=1)
    fees_month = (await db.execute(select(func.sum(FeePayment.amount_paid)).where(FeePayment.school_id == school_id, FeePayment.status == "paid", FeePayment.payment_date >= month_start))).scalar() or 0
    premium = (await db.execute(select(func.count(Student.id)).where(Student.school_id == school_id, Student.is_premium == True, Student.is_active == True))).scalar() or 0

    trend = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        p = (await db.execute(select(func.count(Attendance.id)).where(Attendance.school_id == school_id, Attendance.date == d, Attendance.status == "present"))).scalar() or 0
        t = (await db.execute(select(func.count(Attendance.id)).where(Attendance.school_id == school_id, Attendance.date == d))).scalar() or 0
        trend.append({"date": str(d), "day": d.strftime("%a"), "present": p, "total": t, "percentage": round(p/t*100,1) if t > 0 else 0})

    return {"total_students": total_students, "total_staff": total_staff,
            "today_attendance": {"present": today_present, "total": today_total, "percentage": round(today_present/today_total*100,1) if today_total > 0 else 0},
            "fees_this_month": float(fees_month), "premium_students": premium, "attendance_trend": trend}
